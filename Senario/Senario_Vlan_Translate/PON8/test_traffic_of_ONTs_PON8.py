"""
this trafic checking code is exclusively for ONTs in VOIP Service
"""


import logging
import paramiko
import openpyxl
import os
import socket

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

List_Of_ONTs_On_PON8_Voip_700 = [
     "192.168.120.13",
     "192.168.120.14"
]

List_Of_ONTs_On_PON8_Voip_702 =[
     "192.168.120.17",
     "192.168.120.18",
     "192.168.120.20",
     "192.168.120.21",
     "192.168.120.22"
]


List_Of_ONTs_On_PON8_Voip_802 =[
    "192.168.130.15",
    "192.168.130.16"
]

def is_ssh_connected(ssh):
        if ssh.get_transport():
            transport = ssh.get_transport()
            if transport.is_authenticated():
                return True
        return False

def ssh_connect(IP,username, password):
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(IP, username=username, password=password, port=22, look_for_keys=False, timeout=10)
    except paramiko.AuthenticationException:
        logger.exception("SSH Connect Authentication Failed")
    except socket.timeout:
        logger.exception("SSH Connect Timeout")

    if is_ssh_connected(ssh) == False:
        raise(paramiko.SSHException)
    else:
        return ssh 

def get_information_of_app_on_system(ssh):
    stdin,output,err = ssh.exec_command(f"top -b -n 1", 3)
    app_information = []
    for line in output:
        if line.find("/sp5100/apps/app/i2c")!=-1 or line.find("/sp5100/apps/app/snmp")!=-1 or line.find("/sp5100/apps/app/daemon")!=-1 or line.find("/sp5100/apps/app/maple")!=-1 or line.find("/sp5100/apps/app/katana")!=-1:
            app_information.append(str(line.strip()))
            # logger.info(f"line : {line.strip()}") 
    return app_information

def apply_extract_information(ssh_server,ssh_olt, sheet, workbook, i, List_Of_ONTs):
    logger.info("************************   PON8 ******************************")
    for ont in List_Of_ONTs:
        stdin,output,err = ssh_server.exec_command(f"docker exec freepbx-app{i} ping -c 1 {ont}", 3)
        command = f"docker exec freepbx-app{i} ping -c 1 {ont}"
        logger.info(f"command: {command}")
        stdin_olt,output_olt_time,err_olt = ssh_olt.exec_command(f"uptime", 3)
        for line_olt in output_olt_time:
            time = []
            logger.info(f"line : {line_olt.strip()}")
            time.append(str(line_olt.strip()))
        app_information = get_information_of_app_on_system(ssh_olt)    
        for line in output:
            logger.info(f"line : {line.strip()}")
            if line.find("Request timed out")!=-1 or line.find("Unreachable")!=-1:
                sheet.append([f"{ont}", f"{time[0]}", f"{app_information}"])
                workbook.save("PON8.xlsx")

def test_traffic_of_ONTs():
    #********************************* creat EXCEL report file ************************************
    #current_directory = os.getcwd()
    #logger.info(f"current_directory : {current_directory}")
    workbook = openpyxl.Workbook()
    sheet1 =workbook.create_sheet("PON8 VOIP 700")
    sheet2 =workbook.create_sheet("PON8 VOIP 702")
    sheet3 =workbook.create_sheet("PON8 VOIP 802")
    data = [["ONT IP", "TIME", "INFORMATOPN"]]
    sheet1.append(data[0])
    sheet2.append(data[0])
    sheet3.append(data[0])

    while True:
        try:
            ssh_client_server = ssh_connect("192.168.2.218","saat","1234")
            ssh_client_olt = ssh_connect("192.168.9.135","root","sbkt4v")
            apply_extract_information(ssh_client_server,ssh_client_olt,sheet1,workbook, "", List_Of_ONTs_On_PON8_Voip_700)
            apply_extract_information(ssh_client_server,ssh_client_olt,sheet2,workbook, "2", List_Of_ONTs_On_PON8_Voip_702)
            apply_extract_information(ssh_client_server,ssh_client_olt,sheet3,workbook, "3", List_Of_ONTs_On_PON8_Voip_802)
            ssh_client_server.close()
            ssh_client_olt.close()
        except:
            ssh_client_server.close()
            ssh_client_olt.close()
            continue


def main():
    while True:
        try:
            test_traffic_of_ONTs()
        except:
            continue

if __name__== "__main__" :
    main()
      






